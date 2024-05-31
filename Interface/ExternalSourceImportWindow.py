from os.path import basename
from tkinter import END, Misc, Tk, Toplevel
import tkinter.messagebox as msg
import tkinter.filedialog as fd
from tkinter.ttk import Button, Label, Combobox, Treeview, Scrollbar
from openpyxl import load_workbook
from models import DataImport, Template

class ExternalSourceImportWindow(Toplevel):
    def __init__(self, parent: Toplevel | Tk, master: Misc, template: Template) -> None:
        super().__init__(master)
        self.parent = parent
        self.template = template
        self.prepareInterface()
        self.file_path = None

    def prepareInterface(self):
        self.title("Importuj dane")
        
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(3, weight=1)
        
        self.label = Label(self, text="Select an Excel file:")
        self.select_button = Button(self, text="Browse", command=self.browse_file)
        self.combobox_label = Label(self, text="Worksheet names:")
        self.combobox = Combobox(self)
        self.combobox.bind("<<ComboboxSelected>>", self.update_preview)
        
        self.treeview = Treeview(self, show="headings")
        self.treeview_scroll = Scrollbar(self, orient="vertical", command=self.treeview.yview)
        self.treeview.configure(yscrollcommand=self.treeview_scroll.set)

        self.add_button = Button(self, text="Add", command=self.add_data)

        self.label.grid(row=0, column=0, padx=10, pady=10)
        self.select_button.grid(row=0, column=1, padx=10, pady=10)
        self.combobox_label.grid(row=1, column=0, padx=10, pady=10)
        self.combobox.grid(row=1, column=1, padx=10, pady=10, sticky="ew")
        self.treeview.grid(row=2, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")
        self.treeview_scroll.grid(row=2, column=4, sticky="ns")
        self.add_button.grid(row=3, column=0, columnspan=5, padx=10, pady=10)

    def browse_file(self):
        self.file_path = fd.askopenfilename(
            filetypes=[("Excel files", "*.xlsx;*.xlsm"), ("All files", "*.*")]
        )
        
        if self.file_path:
            self.load_worksheets()

    def load_worksheets(self):
        try:
            workbook = load_workbook(self.file_path, read_only=True)
            sheet_names = workbook.sheetnames
            self.combobox['values'] = sheet_names
            if sheet_names:
                self.combobox.current(0)
                self.update_preview()
        except Exception as e:
            msg.showerror("Error", f"Failed to read the Excel file: {e}")

    def update_preview(self, event=None):
        selected_sheet = self.combobox.get()
        if not selected_sheet or not self.file_path:
            return
        
        try:
            workbook = load_workbook(self.file_path, read_only=True)
            sheet = workbook[selected_sheet]
            
            self.treeview.delete(*self.treeview.get_children())
            first_row = next(sheet.iter_rows(values_only=True))
            if "Email" not in first_row:
                # TODO: Można zrobić jakiś label zamiast treeview i errora
                raise ValueError("Arkusz musi mieć kolumnę 'Email', aby dało się go połączyć z danymi")
                
            self.treeview["columns"] = first_row
            for col in first_row:
                self.treeview.heading(col, text=col)
                self.treeview.column(col, width=100)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.treeview.insert("", END, values=row)
        except Exception as e:
            msg.showerror("Error", f"Failed to read the selected worksheet: {e}")

    def add_data(self):
        di = DataImport(_name=basename(self.file_path), _localPath=self.file_path)
        self.template.dataimport = di
        self.parent.update()
        self.destroy()
