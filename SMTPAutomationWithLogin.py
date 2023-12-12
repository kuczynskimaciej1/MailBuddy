import smtplib
import imaplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import openpyxl

def prepareRecipients():
    RECIPIENTS = []
    CONTACT = []
    SHEET_PATH = input("Enter path to your Excel workbook (example: C:\Dokumenty\Example.xlsx): ")
    WORKBOOK = openpyxl.load_workbook(SHEET_PATH)
    WORKSHEET = WORKBOOK.active
    for ROW in range(0, WORKSHEET.max_row):
        for COLUMN in WORKSHEET.iter_cols(0, WORKSHEET.max_column):
            CELL = COLUMN[ROW].value
            CELL = modifyCell(CELL)
            CONTACT.append(CELL)
        RECIPIENTS.append(CONTACT)
        CONTACT = []
    return RECIPIENTS

def modifyCell(CELL):
    CELL = CELL.replace("ą", "&#261;")
    CELL = CELL.replace("ć", "&#263;")
    CELL = CELL.replace("ę", "&#281;")
    CELL = CELL.replace("ł", "&#322;")
    CELL = CELL.replace("ń", "&#324;")
    CELL = CELL.replace("ś", "&#347;")
    CELL = CELL.replace("ź", "&#378;")
    CELL = CELL.replace("ż", "&#380;")
    CELL = CELL.replace("å", "&aring;")
    CELL = CELL.replace("ä", "&auml;")
    CELL = CELL.replace("ö", "&ouml;")
    CELL = CELL.replace("ø", "&oslash;")
    CELL = CELL.replace("æ", "&aelig;")
    CELL = CELL.replace("þ", "&thorn;")
    CELL = CELL.replace("á", "&aacute;")
    CELL = CELL.replace("ð", "&eth;")
    CELL = CELL.replace("é", "&eacute;")
    CELL = CELL.replace("í", "&iacute;")
    CELL = CELL.replace("ó", "&oacute;")
    CELL = CELL.replace("ú", "&uacute;")
    CELL = CELL.replace("ý", "&yacute;")
    CELL = CELL.replace("Ą", "&#260;")
    CELL = CELL.replace("Ć", "&#262;")
    CELL = CELL.replace("Ę", "&#280;")
    CELL = CELL.replace("Ł", "&#321;")
    CELL = CELL.replace("Ń", "&#323;")
    CELL = CELL.replace("Ś", "&#346;")
    CELL = CELL.replace("Ź", "&#377;")
    CELL = CELL.replace("Ż", "&#379;")
    CELL = CELL.replace("Å", "&Aring;")
    CELL = CELL.replace("Ä", "&Auml;")
    CELL = CELL.replace("Ö", "&Ouml;")
    CELL = CELL.replace("Ø", "&Oslash;")
    CELL = CELL.replace("Æ", "&AElig;")
    CELL = CELL.replace("Þ", "&THORN;")
    CELL = CELL.replace("Á", "&Aacute;")
    CELL = CELL.replace("Ð", "&ETH;")
    CELL = CELL.replace("É", "&Eacute;")
    CELL = CELL.replace("Í", "&Iacute;")
    CELL = CELL.replace("Ó", "&Oacute;")
    CELL = CELL.replace("Ú", "&Uacute;")
    CELL = CELL.replace("Ý", "&Yacute;")
    return CELL

def prepareMail(FROM, RECIPIENT, NAME, JOB, ATTACHMENT_PATH):
    MESSAGE = MIMEMultipart('mixed')
    MESSAGE['Subject'] = "JoinThe.Space - networking offer"
    MESSAGE['From'] = FROM
    MESSAGE['To'] = RECIPIENT[0]
    HTML = """\
    <html>
      <head></head>
      <body>
        <p>Cześć!<br><br>Ruszamy z kolejnymi kosmicznymi projektami, tym razem docierając bezpośrednio na Uczelnie!
        <br><br>
        Chcielibyśmy Was zaprosić na <strong>jesienne spotkanie</strong>, które przeprowadzilibyśmy osobiście na {0} w ramach trasy <strong>Space Talks: How to JoinThe.Space Industry?</strong>. Projekt chcemy przeprwadzić pod patronatem świetnych organizacji, w tym Polskiej Agencji Kosmicznej, i chcemy dotrzeć do wszystkich zaprzyjaźnionych uczelni w całej Europie
        <br><br>
        Poznamy się bliżej, poznacie naszą działalność oraz perspektywy rozwoju na rynku kosmicznym, które czekają na studentów Waszej uczelni, takie jak staże i praktyki w firmach z branży kosmicznej.
        Chcielibyśmy wiedzieć, czy jesteście zainteresowani tą inicjatywą? Jeśli tak - umówmy się na spotkanie na Google Meets w dogodnym terminie, aby omówić szczegóły organizacji, takie jak podział zadań, data i lokalizacja.
           <div id="isPasted">Best regards,</div>
    <div><strong>{1}</strong></div>
    <div>{2}</div>
    <div><a href="mailto:{3}" rel="noopener noreferrer" target="_blank">{4}</a></div>
    <div>
	    <br>
    </div>
    <div>JoinThe.Space</div>
    <div><a href="https://www.jointhe.space" rel="noopener noreferrer" target="_blank">www.jointhe.space</a></div>
    <div>
	    <br>
    </div>
    <div><i style="font-family: Arial, Helvetica, sans-serif; font-size: small; font-variant-ligatures: normal; font-variant-caps: normal; font-weight: 400; 
    letter-spacing: normal; orphans: 2; text-align: start; text-indent: 0px; text-transform: none; white-space: normal; widows: 2; word-spacing: 0px; -webkit-text-stroke-width: 0px; 
    background-color: rgb(255, 255, 255); text-decoration-thickness: initial; text-decoration-style: initial; text-decoration-color: initial; color: rgb(0, 0, 0);">
    <span style="font-size: 8pt;">This email and its attachments contain information that is confidential, proprietary, and only for the use of the intended recipient. If you are not the intended recipient, please notify us and do not copy or forward this email or its attachments.</span></i><span style="font-size: 8pt;">&nbsp;</span></div>
        </p>
      </body>
    </html>
    """.format(RECIPIENT[1], NAME, JOB, FROM, FROM)
    TEXT = MIMEText(HTML, 'html')
    ATTACHMENT = loadAttachment(ATTACHMENT_PATH)
    MESSAGE.attach(TEXT)
    MESSAGE.attach(ATTACHMENT)
    return MESSAGE.as_string()

def loadAttachment(ATTACHMENT_PATH):
    ATTACHMENT = MIMEApplication(open(ATTACHMENT_PATH, "rb").read(),_subtype="pdf")
    ATTACHMENT.add_header('Content-Disposition', "attachment; filename= %s" %
    ATTACHMENT_PATH.split("\\")[-1])
    return ATTACHMENT

def sendMail(FROM, PASSWORD, TO, MESSAGE):
    SMTP_HOST = "smtp.titan.email"
    SMTP_PORT = 465
    SERVER = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT)
    SERVER.connect(SMTP_HOST, SMTP_PORT)
    SERVER.ehlo()
    SERVER.login(FROM, PASSWORD)
    SERVER.sendmail(FROM, TO, MESSAGE)
    SERVER.quit()

def saveInSent(FROM, PASSWORD, MESSAGE):
    IMAP_HOST = "imap.titan.email"
    IMAP_PORT = 993
    IMAP = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    IMAP.login(FROM, PASSWORD)
    IMAP.append("Sent", "\\Seen", imaplib.Time2Internaldate(time.time()), MESSAGE.encode('utf8'))
    IMAP.logout()

FROM = input("Enter your e-mail address: ")
PASSWORD = input("Enter your e-mail password: ")
NAME = input("Enter your name (eg. Jan Kowalski): ")
NAME = modifyCell(NAME)
JOB = input("Enter your function (eg. Head of Western Poland): ")
ATTACHMENT_PATH = input("Enter path to your attachment (example: C:\Dokumenty\ExamplePDF.pdf): ")
RECIPIENTS = prepareRecipients()

for RECIPIENT in RECIPIENTS:
    MESSAGE = prepareMail(FROM, RECIPIENT, NAME, JOB, ATTACHMENT_PATH)
    sendMail(FROM, PASSWORD, RECIPIENT, MESSAGE)
    print("Sent to: " + str(RECIPIENT))
    saveInSent(FROM, PASSWORD, MESSAGE)
    print("Saved in sent.")
    print()