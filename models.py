from __future__ import annotations
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
from xml.etree.ElementTree import ParseError
from sqlalchemy import BOOLEAN, Column, ForeignKey, Integer, String, LargeBinary, TIMESTAMP, func
from sqlalchemy.orm import declarative_base, relationship
from sqlalchemy.ext.hybrid import hybrid_property
import re
from dns.resolver import resolve
import requests
import smtplib
import imaplib
import xml.etree.ElementTree as ET


__all__ = ["Template", "Attachment", "Contact", "User", "Message", "Group"]


class IModel(declarative_base()):
    __abstract__ = True
    run_loading = True
    addQueued: list[IModel] = []
    updateQueued: list[IModel] = []
    retrieveAdditionalQueued: list[IModel] = []

    @staticmethod
    def queueSave(child):
        if not IModel.run_loading:
            IModel.addQueued.append(child)
    
    @staticmethod
    def queueToUpdate(child):
        if not IModel.run_loading:
            IModel.updateQueued.append(child)
       
    @staticmethod
    def retrieveAdditionalData(child):
        if isinstance(child, Template):
            IModel.retrieveAdditionalQueued.append(child)


class DataImport(IModel):
    all_instances: list[DataImport] = []
    __tablename__ = "DataImport"

    _id = Column("id", Integer, primary_key=True, autoincrement=True)
    _name = Column("name", String(100))
    _localPath = Column("localPath", String(255), nullable=True)
    _content = Column("content", LargeBinary, nullable=True)
    
    def __init__(self, **kwargs) -> None:
        self.id = kwargs.pop('_id', None)
        self.name = kwargs.pop('_name', None)
        self.localPath = kwargs.pop('_localPath', None)
        self.content = kwargs.pop('_content', None)
        DataImport.all_instances.append(self)
        IModel.queueSave(child=self)
        
    def getColumnPreview(self) -> dict | None:
        workbook = load_workbook(self.localPath, read_only=True)
        result = dict()
        for sheet in workbook:
            first_row = next(sheet.iter_rows(values_only=True))
            if "Email" not in first_row:
                continue
            
            columns = first_row
            dataPreviewRow = next(sheet.iter_rows(min_row=2, values_only=True))
            for idx, c in enumerate(columns):
                result[c] = dataPreviewRow[idx]
        return result if len(result) > 0 else None
    
    def GetRow(self, email: str) -> dict[str, str]:
        workbook = load_workbook(self.localPath, read_only=True)
        result = dict()
        for sheet in workbook:
            first_row = next(sheet.iter_rows(values_only=True))
            if "Email" not in first_row:
                continue
            emailColumnIdx = first_row.index("Email")
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[emailColumnIdx] == email:
                    for idx, column in enumerate(first_row):
                        result[column] = row[idx]
                        break
                    break
            break
        if len(result) == 0:
            raise AttributeError("Nie znaleziono odpowiadającej linijki w pliku z danymi do uzupełnienia")
        return result
            

# region Properties
    @hybrid_property
    def id(self):
        return self._id

    @hybrid_property
    def name(self):
        return self._name

    @hybrid_property
    def content(self):
        return self._content
    
    @hybrid_property
    def localPath(self):
        return self._localPath
    
    @id.setter
    def id(self, newValue: int):
        self._id = newValue

    @name.setter
    def name(self, value: str | None):
        self._name = value
        IModel.queueToUpdate(self)

    @content.setter
    def content(self, value: object | None):
        self._content = value
        IModel.queueToUpdate(self)
        
    @localPath.setter
    def localPath(self, value: str | None):
        self._localPath = value
        IModel.queueToUpdate(self)
#endregion


class Template(IModel):
    all_instances: list[Template] = []
    __tablename__ = "Templates"

    _id = Column("id", Integer, primary_key=True, autoincrement=True)
    _name = Column("name", String(100), nullable=True)
    _content = Column("content", String, nullable=True)
    _dataimport_id = Column("dataimport_id", Integer, #ForeignKey("DataImport.id", ondelete='SET NULL'), 
                            nullable=True)
    
    # dataImportRel = relationship(DataImport, foreign_keys=[DataImport._id])
    
    def __init__(self, **kwargs) -> None:
        self.id: int = kwargs.pop('_id', None)
        self.name: str = kwargs.pop('_name', None)
        self.content: object = kwargs.pop('_content', None)
        self.dataimport: DataImport = None
        self.dataimport_id: int = kwargs.pop("_dataimport_id", None)
        Template.all_instances.append(self)
        IModel.queueSave(child=self)


    def __str__(self) -> str:
        return self.name if self.name != None else "<puste>"
    
    def __repr__(self):
        return f"Template(_name={self.name}, _content={self.content}, _id={self.id})"
    
    def FillGaps(self, fillData: dict[str, str]) -> str:
        html_text = self.content

        span_text = '<span>' # TODO korzystać z tej metody do każdego generowania podglądu
        pattern = r"<MailBuddyGap>\s*([^<>\s][^<>]*)\s*</MailBuddyGap>"
        matches = re.findall(pattern, html_text)
        
        for m in matches:
            try:
                preview_text = fillData[m]
            except KeyError as ke:
                raise AttributeError("Nie znaleziono odpowiadającej wartości dla luki", ke)
            tmp = html_text.replace(f"<MailBuddyGap>{m}</MailBuddyGap>", span_text + preview_text + "</span>")
            if tmp:
                html_text = tmp
        return html_text

# region Properties
    @hybrid_property
    def id(self):
        return self._id

    @hybrid_property
    def name(self):
        return self._name

    @hybrid_property
    def content(self):
        return self._content
    
    @hybrid_property
    def dataimport_id(self) -> DataImport:
        return self._dataimport_id

    @id.setter
    def id(self, newValue: int):
        # TODO: if initial setup / loading from db
        self._id = newValue

    @name.setter
    def name(self, value: str | None):
        self._name = value
        IModel.queueToUpdate(self)

    @content.setter
    def content(self, value: str | None):
        self._content = value
        IModel.queueToUpdate(self)
        
    @dataimport_id.setter
    def dataimport_id(self, value: int | None):
        self._dataimport_id = value
        IModel.queueToUpdate(self)
        IModel.retrieveAdditionalData(self)
#endregion


class Attachment(IModel):
    all_instances = []
    __tablename__ = "Attachments"

    attachment_id = Column(Integer, primary_key=True)
    name = Column(String(100), nullable=False)
    file_path = Column(String(255), nullable=True)
    file = Column(LargeBinary, nullable=True)

    def __init__(self, path, type, attachment_id: int | None = None) -> None:
        self.attachment_id = attachment_id
        self.path = path
        self.type = type
        Attachment.all_instances.append(self)
        IModel.queueSave(child=self)

    # def prepareAttachment(self):
    #     att = MIMEApplication(open(self.path, "rb").read(), _subtype=self.type)
    #     att.add_header('Content-Disposition',
    #                    "attachment; filename= %s" % self.path.split("\\")[-1])
    #     return att


class Contact(IModel):
    all_instances = []
    __tablename__ = "Contacts"

    _email = Column("email", String(100), primary_key=True)
    _first_name = Column("first_name", String(50))
    _last_name = Column("last_name", String(50))

    def __init__(self, **kwargs) -> None:
        """Creates instance and adds it to all_instances
        Args:
            first_name (str): any string
            last_name (str): any string
            email (str): must match standard pattern x@y.z
        Raises:
            AttributeError: when email doesn't match standard pattern
        """
        self.email = kwargs.pop("_email", None)
        self.first_name = kwargs.pop("_first_name", "")
        self.last_name = kwargs.pop("_last_name", "")
        Contact.all_instances.append(self)
        IModel.queueSave(child=self)

    def __str__(self) -> str:
        return f"{self.first_name} {self.last_name}, <{self.email}>"

    def __eq__(self, other) -> bool:
        if not isinstance(other, Contact):
            return NotImplemented
        return self.email == other.email and self.first_name == other.first_name and self.last_name == other.last_name

    @classmethod
    def get_by_id(cls, searched_id: int) -> Contact | None:
        for candidate in cls.all_instances:
            if candidate.id == searched_id:
                return candidate
        return None

    @staticmethod
    def isEmail(candidate: str) -> bool:
        if candidate == None:
            return False
        
        if re.match(r"^(?!.*@.*@.*$)[^@]+@[^@]+\.[^@]+$/g", candidate):
            return False    
        
        return True
    
# region Properties
    @hybrid_property
    def email(self):
        return self._email

    @hybrid_property
    def first_name(self):
        return self._first_name
    
    @hybrid_property
    def last_name(self):
        return self._last_name

    @email.setter
    def email(self, newValue: str):
        if not Contact.isEmail(newValue):
            raise AttributeError(f"{newValue} is not an email")
        self._email = newValue

    @first_name.setter
    def first_name(self, value: str | None):
        self._first_name = value

    @last_name.setter
    def last_name(self, value: str | None):
        self._last_name = value
#endregion


class User(IModel):
    all_instances = []
    __tablename__ = "Users"

    _id = Column("_id", Integer, primary_key=True, autoincrement=True)
    _email = Column("email", String(100), ForeignKey('Contacts.email'), unique=True)
    _selected = Column("selected", BOOLEAN)
    
    contactRel = relationship(Contact, foreign_keys=[_email])

    default_settings = {
    'gmail.com': {
        'imap': {'hostname': 'imap.gmail.com', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.gmail.com', 'port': 465, 'socket_type': 'SSL'}
    },
    'yahoo.com': {
        'imap': {'hostname': 'imap.mail.yahoo.com', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.mail.yahoo.com', 'port': 465, 'socket_type': 'SSL'}
    },
    'outlook.com': {
        'imap': {'hostname': 'outlook.office365.com', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.office365.com', 'port': 587, 'socket_type': 'STARTTLS'}
    },
    'poczta.onet.pl': {
        'imap': {'hostname': 'imap.poczta.onet.pl', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.poczta.onet.pl', 'port': 465, 'socket_type': 'SSL'}
    },
    'onet.pl': {
        'imap': {'hostname': 'imap.poczta.onet.pl', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.poczta.onet.pl', 'port': 465, 'socket_type': 'SSL'}
    },
    'wp.pl': {
        'imap': {'hostname': 'imap.wp.pl', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.wp.pl', 'port': 465, 'socket_type': 'SSL'}
    },
    'interia.pl': {
        'imap': {'hostname': 'imap.poczta.interia.pl', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.poczta.interia.pl', 'port': 465, 'socket_type': 'SSL'}
    },
    'pcz.pl': {
        'imap': {'hostname': 'imap.pcz.pl', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.pcz.pl', 'port': 465, 'socket_type': 'SSL'}
    },
    'wimii.pcz.pl': {
        'imap': {'hostname': 'imap.wimii.pcz.pl', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.wimii.pcz.pl', 'port': 465, 'socket_type': 'SSL'}
    },
    
    'ethereal.email': {
        'imap': {'hostname': 'imap.ethereal.email', 'port': 993, 'socket_type': 'SSL'},
        'smtp': {'hostname': 'smtp.ethereal.email', 'port': 587, 'socket_type': 'STARTTLS'}
    }
}
    
    def __init__(self, **kwargs) -> None:
        self.email = kwargs.pop("_email")
        self.password = kwargs.pop("_password", None)
        self.contact = self.getExistingContact(kwargs.pop("_first_name", None), kwargs.pop("_last_name", None))
        self.selected = kwargs.pop("_selected", False)
        self._smtp_host = ""
        self._smtp_port = ""
        self._smtp_socket_type = "SSL"
        User.all_instances.append(self)
        IModel.queueSave(child=self)
        
# region Properties
    @hybrid_property
    def email(self):
        return self._email

    @email.setter
    def email(self, newValue: str):
        self._email = newValue
        
    @hybrid_property
    def selected(self) -> bool:
        return self._selected

    @selected.setter
    def selected(self, newValue: bool):
        if newValue == True:
            for u in User.all_instances:
                u._selected = False
        self._selected = newValue
#endregion

    @staticmethod
    def get_domain(email):
        return email.split('@')[1]

    @staticmethod
    def get_mx_records(domain):
        try:
            answers = resolve(domain, 'MX')
            mx_records = [answer.exchange.to_text() for answer in answers]
            return mx_records
        except Exception as e:
            print(f"DNS lookup failed: {e}")
            return []


    @staticmethod
    def get_autodiscover_settings(domain):
        try:
            url = f'https://autoconfig.{domain}/mail/config-v1.1.xml'
            response = requests.get(url)
            if response.status_code == 200:
                return response.text
            else:
                url = f'https://{domain}/.well-known/autoconfig/mail/config-v1.1.xml'
                response = requests.get(url)
                if response.status_code == 200:
                    return response.text
        except Exception as e:
            print(f"Autodiscover failed: {e}")
        return None


    @staticmethod
    def parse_email_settings(xml_data):
        try:
            tree = ET.ElementTree(ET.fromstring(xml_data))
        except ParseError:
            raise AttributeError("Zwrócono niepoprawny settings XML")
        
        root = tree.getroot()
        email_provider = root.find('emailProvider')

        settings = {
            'imap': {},
            'pop3': {},
            'smtp': []
        }

        for server in email_provider.findall('incomingServer'):
            server_type = server.get('type')
            settings[server_type] = {
                'hostname': server.find('hostname').text,
                'port': int(server.find('port').text),
                'socket_type': server.find('socketType').text
            }

        for server in email_provider.findall('outgoingServer'):
            smtp_settings = {
                'hostname': server.find('hostname').text,
                'port': int(server.find('port').text),
                'socket_type': server.find('socketType').text
            }
            settings['smtp'].append(smtp_settings)

        return settings
    

    def test_imap_connection(imap_settings, email, password):
        try:
            if imap_settings == 'SSL':
                connection = imaplib.IMAP4_SSL(imap_settings['hostname'], imap_settings['port'])
            else:
                connection = imaplib.IMAP4(imap_settings['hostname'], imap_settings['port'])
            
            connection.login(email, password)
            connection.logout()
            return True
        except Exception as e:
            print(f"IMAP connection failed: {e}")
            return False


    def test_smtp_connection(smtp_settings, email, password):
            try:
                if smtp_settings == 'SSL':
                    connection = smtplib.SMTP_SSL(smtp_settings['hostname'], smtp_settings['port'])
                else:
                    connection = smtplib.SMTP(smtp_settings['hostname'], smtp_settings['port'])
                    if smtp_settings == 'STARTTLS':
                        connection.starttls()

                connection.login(email, password)
                connection.quit()
                return True
            except Exception as e:
                print(f"SMTP connection to {smtp_settings['hostname']} on port {smtp_settings['port']} failed: {e}")
            return False
    

    def discover_email_settings(self):
        domain = User.get_domain(self.email)
        settings_xml = User.get_autodiscover_settings(domain)
        if not settings_xml:
            raise AttributeError("Nie znaleziono opcji dla podanej domeny")
        
        # try:
        settings_xml = User.parse_email_settings(settings_xml)
        # except 
        
        if domain in self.default_settings:
            settings_xml = self.default_settings[domain]
            print(settings_xml)
            return settings_xml
        
        if self.test_imap_connection(settings_xml['imap'], self.email, self.password)  \
            and self.test_smtp_connection(settings_xml['smtp'], self.email, self.password):
                print("Check ok")
                print(settings_xml)
                return settings_xml
        
        raise AttributeError("Nie znaleziono opcji dla podanej domeny")
    
    
    @staticmethod
    def GetCurrentUser() -> User | None:
        for u in User.all_instances:
            if u._selected:
                return u
        return None
        
    def getExistingContact(self, first_name, last_name) -> Contact:
        for c in Contact.all_instances:
            if c.email == self._email:
                return c
        return Contact(_first_name=first_name, _last_name=last_name, _email=self._email)


class Message(IModel):
    all_instances = []
    __tablename__ = "Messages"

    message_id = Column(Integer, primary_key=True)
    trigger_id = Column(Integer, nullable=False)
    email = Column(String(100), nullable=False)
    template_id = Column(Integer, nullable=False)
    sent_at = Column(TIMESTAMP, default=func.now())

    # TODO
    # trigger = relationship("Trigger")
    # contact = relationship("Contact")
    # template = relationship("Template")

    def __init__(self, template: Template,  recipient: Contact,
                 att: list[Attachment] = []) -> None:
        self.recipient = recipient
        self.email = recipient.email
        self.att = att
        self.template = template
        self.template_id = template.id
        Message.all_instances.append(self)
        IModel.queueSave(child=self)
        
    def getParsedBody(self) -> str:
        data = dict()
        if self.template.dataimport:
            data = self.template.dataimport.GetRow(self.email)
        body: str = self.template.FillGaps(data)
        return body
    


class Group(IModel):
    all_instances: list[Group] = []
    __tablename__ = "Groups"
    
    _id = Column("id", Integer, primary_key=True)
    _name = Column("name", String(100), nullable=True)
    
    def __init__(self, **kwargs):
        self.id: int = kwargs.pop('_id', None)
        self.name: str = kwargs.pop('_name', "")
        self.contacts: list[Contact] = kwargs.pop("_contacts", [])
        Group.all_instances.append(self)
        IModel.queueSave(self)
        
    @hybrid_property
    def name(self):
        return self._name

    @name.setter
    def name(self, value: str | None):
        self._name = value
        IModel.queueToUpdate(self)
    
    def __str__(self):
        return f"{self.id}: {self.name}"
    
    @classmethod
    def get_by_id(cls, searched_id: int) -> Group | None:
        for candidate in cls.all_instances:
            if candidate.id == searched_id:
                return candidate
        return None
    
    def _add_contact(self, c: Contact) -> bool:
        if c not in self.contacts:
            self.contacts.append(c)
            return True
        return False
    
#region Properties
    @hybrid_property
    def id(self):
        return self._id

    @hybrid_property
    def name(self):
        return self._name
    
    @id.setter
    def id(self, newValue: int):
        if newValue:
            self._id = newValue
        else:
            self._id = max((i.id for i in Group.all_instances), default=-1) + 1

    @name.setter
    def name(self, value: str | None):
        self._name = value
        IModel.queueToUpdate(self)
#endregion
